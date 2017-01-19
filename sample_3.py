#Worksheet in TIMESHEET PROCESS -NON IBM.xlsx
#both the modules openpyxl and xlrd have some roles in the following code
import openpyxl
import xlrd
import re
#------------------------------------------------------------------------------#

#------------------------------------------------------------------------------#
#the function get_row_column returns the row and column number of a given value
def get_row_column(worksheet,value):
    for row in range(worksheet.nrows):
        for column in range(worksheet.ncols):
            if worksheet.cell(row,column).value == value:
                return row,column
#------------------------------------------------------------------------------#

#------------------------------------------------------------------------------#
def excel_extract(path):
    #xlrd.open_workbook(path,'cp1252') takes care of the unicodes
    workbook = xlrd.open_workbook(path,
                                  'cp1252')
    worksheet = workbook.sheet_by_index(0)
    r1,c1 = get_row_column(worksheet,"Vendor company")
    #print r1,c1
    months = ["January","February","March","April","May","June","July","August",
              "September","October","November","December"]
    #print str((worksheet.cell(0,0).value))
    d = re.findall("\d+",str((worksheet.cell(0,0).value)))
    m = []
    index = []
    i = 0
    for month in months:
        if re.search(month,str((worksheet.cell(0,0).value))):
            index.append(i)
            m.append(re.findall(month,str((worksheet.cell(0,0).value))))
        i = i + 1
    #print m
    #print index
    if index[0]%2 == 0:
        row_num = 31 + (int(d[1]) - int(d[0]))
    elif index[0] == 1:
        row_num = 28 + (int(d[1]) - int(d[0]))
    elif index[0]%2 == 1:
        row_num = 30 + (int(d[1]) - int(d[0]))
    #print row_num
    column_head = []
    for j in range(c1,c1 + row_num + 8,1):
        column_head.append(str((worksheet.cell(r1,j).value)))
    #print column_head
    data_list = []
    i = 0
    while str((worksheet.cell(r1 + 1 +i,0).value)) != '':
        
        for j in range(c1,c1 + row_num + 8,1):
            data_list.append(str((worksheet.cell(r1 + 1 +i,j).value)))
        
        i = i + 1
    #print data_list
    c = []
    i = 0
    days_list = []
    for j in range(c1 + 4,c1 + row_num + 8,1):
        days_list.append(str((worksheet.cell(r1 - 1,j).value)))
    #print days_list
    #print column_head[6:]
    #print data_list[6:]
    m_file = zip(column_head[6:],days_list[2:],data_list[6:])
    #print m_file
    main_dict_list = []
    i = 0
    for m in m_file:
        
        main_dict = {'dates':m_file[i][0],
                     'days':m_file[i][1],
                     'attendence':m_file[i][2]}
        main_dict_list.append(main_dict)
        i = i + 1
    i = 0
    table_main_dict = {}
    for c in column_head[:5]:
        table_main_dict.update({
            (column_head[:5])[i]:(data_list[:5])[i]})
        i = i + 1
    #print main_dict_list
    file1 = open("sample_3.txt",'w')
    for dictionary in main_dict_list:
        for key,value in dictionary.iteritems():
            print key,value
            file1.write(key + ":"  +  value)
            file1.write("\t")
        file1.write("\n")
        file1.write("\n")
    for key,value in table_main_dict.iteritems():
        print key,value
        file1.write(key + ":"  +  value)
        file1.write("\t")
    file1.write("\n")
    file1.close()
    return main_dict_list,table_main_dict
a,b = excel_extract("sample_3.xlsx")
print a,b
        
    
 
