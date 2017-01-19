import xlrd
import openpyxl
import re
#this is the program for the extraction of sample_1.xlsx
#------------------------------------------------------------------------------#

#------------------------------------------------------------------------------#
def get_row_column(path,value):
    workbook = xlrd.open_workbook(path,'cp1252')
    worksheet = workbook.sheet_by_index(0)
    sheet = workbook.sheet_by_index(0)
    for row in range(sheet.nrows):
        for column in range(sheet.ncols):
            if sheet.cell(row,column).value == value:
                return row,column
#------------------------------------------------------------------------------#

#------------------------------------------------------------------------------#
def excel_extract(path):
    #xlrd.open_workbook(path,'cp1252') takes care of the unicodes
    workbook = xlrd.open_workbook(path,'cp1252')
    worksheet = workbook.sheet_by_index(0)
    list_index = []
    particulars_list = []
    particulars_dict_list = []
    sheet = workbook.sheet_by_index(0)
    lr = sheet.nrows
    print("no of cols: %d" %sheet.ncols)

    r1,c1 = get_row_column(path,'ADDRESS')
    r1 = r1 - 1
    wr,wc = get_row_column(path,'WORDS')
    main_table_r,main_table_c = get_row_column(path,'S No')
    print main_table_r,main_table_c
    print wr,wc
    wb = openpyxl.load_workbook(path,'cp1252')
    sheets = wb.get_sheet_names()
    totals = []
    a = []
    table_part_1 = []
    for k in range(wr + 1,lr,1):
        table_part_1.append(str((worksheet.cell(k,wc + 0).value).encode('ascii','ignore')))
        if re.search(':' or '|' ,table_part_1[k-wr - 1]):
            table_part_1[k-wr - 1] = str((worksheet.cell(k,wc + 0).value).encode('ascii','ignore')) + str((worksheet.cell(k,wc + 1).value))
    print table_part_1
    particulars = []
    heading = ''
    counter1 = 0
    for k in range(main_table_r + 1,wr,1):
        particulars.append(str((worksheet.cell(k,1).value).encode('ascii','ignore')))
        if (re.search(':',particulars[k-main_table_r - 1]) == None)and(particulars[k-main_table_r - 1] != ''):
            heading = particulars[k-main_table_r - 1]
        if re.search(':',particulars[k-main_table_r - 1]):       
            particulars[k-main_table_r - 1] = str((worksheet.cell(k,main_table_c +1).value).encode('ascii','ignore')) + str((worksheet.cell(k,2).value))
            if (re.search(':',particulars[k-main_table_r - 1]))and(particulars[k-main_table_r - 1] != ''):
                counter1 = 1
                particulars_list.append(particulars[k-main_table_r - 1])
        if (re.search(':',particulars[k-main_table_r - 1]) == None)and(counter1 == 1):
            list_index.append(k)
            dict1 = {heading : particulars_list}
            particulars_dict_list.append(dict1)
            particulars_list = []
            counter1 = 0
    print particulars
    print particulars_list
    print particulars_dict_list
    billable_days = []
    for k in range(main_table_r + 1,wr,1):
        billable_days.append(str((worksheet.cell(k,3).value)))
        if re.search(':',billable_days[k-11]):
            billable_days[k-11] = str((worksheet.cell(k,main_table_c +3).value)) + str((worksheet.cell(k,4).value))
    print "billable days:",billable_days
    billable_qty = []
    for k in range(main_table_r + 1,wr,1):
        billable_qty.append(str((worksheet.cell(k,main_table_c +4).value)))
       
    print "billable quantity:",billable_qty
    unit_price = []
    for k in range(main_table_r + 1,wr,1):
        unit_price.append(str((worksheet.cell(k,main_table_c +5).value)))
       
    print "unit price:",unit_price
    amount = []
    for k in range(main_table_r + 1,wr,1):
        amount.append(str((worksheet.cell(k,main_table_c +6).value)))
       
    print amount
    s_no = []
    for k in range(main_table_r + 1,wr,1):
        s_no.append(str((worksheet.cell(k,0).value)))
        if re.search('Total',s_no[k-main_table_r - 1]):
            totals.append(str((worksheet.cell(k,main_table_c +0).value)) +":" +  str((worksheet.cell(k,main_table_c +6).value)))
       
    print "s_no:",s_no
    print totals
    print list_index


    list_dict = []
    for index in list_index:
        dict1 = {str((worksheet.cell(main_table_r,main_table_c +0).value)):s_no[index - main_table_r - 5],
                 str((worksheet.cell(main_table_r,main_table_c +1).value)):particulars[index - main_table_r - 5],
                 str((worksheet.cell(main_table_r,main_table_c +3).value)):billable_days[index - main_table_r - 5],
                 str((worksheet.cell(main_table_r,main_table_c +4).value)):billable_qty[index - main_table_r - 5],
                 str((worksheet.cell(main_table_r,main_table_c +5).value)):unit_price[index - main_table_r - 5],
                 str((worksheet.cell(main_table_r,main_table_c +6).value)):amount[index  - main_table_r - 5]
            }
        list_dict.append(dict1)
    print list_dict
    for i in sheets:
        sheet = wb.get_sheet_by_name(i)
        for j in range(5,11,1):
            if sheet.cell(row = j,column = 1).value:
                if is_ascii(sheet.cell(row = j,column = 1).value):
                    
                    
                    a.append(str((sheet.cell(row = j,column = 1).value)
                             .encode('ascii','ignore')))
                else:
                    a.append(str((sheet.cell(row = j,column = 1).value)))
                    
            else:
                a.append(None)
        print a
    for j in range(7,11,1):
        if sheet.cell(row = j,column  = 6).value:
            if is_ascii(sheet.cell(row = j,column = 6).value):
            
                a.append(str((sheet.cell(row = j,column = 6).value)
                         .encode('ascii','ignore')))
            else:
                a.append(str((sheet.cell(row = j,column = 6).value)))
                
            
        else:
            a.append(None)
    print a
    introduction = []
    for k in range(r1 - 1,main_table_r ,1):
        introduction.append(str((worksheet.cell(k,c1).value)))
       
    print introduction
    file1 = open("sample_1.txt",'w')
    for dictionary in particulars_dict_list:
        for key,value in dictionary.iteritems():
            print key,value
            file1.write(key + ":" )
            for i in value:
                 file1.write(i)
                 file1.write("\t")
            file1.write("\t")
        file1.write("\n")
        file1.write("\n")
    for dictionary in list_dict:
        for key,value in dictionary.iteritems():
            print key,value
            file1.write(key + ":"  +  value)
            file1.write("\t")
        file1.write("\n")
        file1.write("\n")
    for i in introduction:
        file1.write(i)
        file1.write("\n")
    file1.write("\n")
    for i in introduction:
        file1.write(i)
        file1.write("\t")
    file1.write("\n")
    print particulars_dict_list,list_dict,introduction,totals
    return particulars_dict_list,list_dict,introduction,totals
def is_ascii(s):
    return all(ord(c) < 128 for c in s)
excel_extract("sample_1.xlsx")
