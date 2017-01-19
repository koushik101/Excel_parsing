#both the modules openpyxl and xlrd have some roles in the following code
import openpyxl
import xlrd
#the function get_row_column returns the row and column number of a given value
#------------------------------------------------------------------------------#

#------------------------------------------------------------------------------#
def get_value_column(worksheet,row_num ,column_num):
    for j in range(int(column_num) + 1,int(column_num) + 20,1):
        
        if str((worksheet.cell(int(row_num),j).value)) != '':
            return j
        else:
            return 10
#------------------------------------------------------------------------------#

#------------------------------------------------------------------------------#
def get_row_column(value,worksheet):
    for row in range(worksheet.nrows):
        for column in range(worksheet.ncols):
            if worksheet.cell(row,column).value == value:
                return row,column
#------------------------------------------------------------------------------#

#------------------------------------------------------------------------------#
def excel_extract(path):
    #xlrd.open_workbook(path,'cp1252') takes care of the unicodes
    workbook = xlrd.open_workbook(path,'cp1252')
    worksheet = workbook.sheet_by_index(0)
    wb = openpyxl.load_workbook(path,'cp1252')
    sheets = wb.get_sheet_names()
    #print sheets
    a = []
    b = []
    lr = worksheet.nrows

    r1,c1 = get_row_column('Year',worksheet)
    r2,c2 = get_row_column('S.No.',worksheet)
    r3,c3 = get_row_column('Approved by:',worksheet)
    r4,c4 = get_row_column('Contractor Signature:',worksheet)
    cosultantr,consultantc = get_row_column('Consultant',worksheet)
    #print cosultantr,consultantc
    #print r4,c4
    #print r1,c1
    #print r2,c2
    for i in sheets:
        sheet = wb.get_sheet_by_name(i)
        for j in range(r1,r2,1):
            a.append(str((worksheet.cell(j,c1 + 0).value).encode('ascii','ignore')))
        #print a
        for j in range(r1,r2,1):
            a.append(str((worksheet.cell(j,c1 + 5).value).encode('ascii','ignore')))
        #print a
        for j in range(r1,r2,1):
            b.append(str((worksheet.cell(j,c1 + 2).value)))
        #print b
        for j in range(r1,r2,1):
            b.append(str((worksheet.cell(j,c1 + 8).value)))
        #print b
 
    i = 0
    dict2 = {} 
    for an in a:
        dict2.update({a[i]:b[i]})
        i = i + 1 
    #print dict2
    s_no = []
    for k in range(r2 + 1,r3,1):
        s_no.append(str((worksheet.cell(k,c2 + 0).value)))
    #print s_no
    effort_date = []
    for k in range(r2 + 1,r3,1):
        effort_date.append(str((worksheet.cell(k,c2 +  1).value)))
    #print effort_date
    Day = []
    for k in range(r2 + 1,r3,1):
        Day.append(str((worksheet.cell(k,c2 + 2).value)))
    #print Day
    in_time = []
    for k in range(r2 + 1,r3,1):
        in_time.append(str((worksheet.cell(k,c2 + 3).value)))
    #print in_time
    out_time = []
    for k in range(r2 + 1,r3,1):
        out_time.append(str((worksheet.cell(k,c2 + 4).value)))
    #print out_time
    attendence = []
    for k in range(r2 + 1,r3,1):
        attendence.append(str((worksheet.cell(k,c2 + 5).value)))
    #print attendence
    hours = []
    for k in range(r2 + 1,r3,1):
        hours.append(str((worksheet.cell(k,c2 +7).value)))
    #print hours
    project_id = []
    for k in range(r2 + 1,r3,1):
        project_id.append(str((worksheet.cell(k,c2 + 8).value)))
    #print project_id
    m_file = zip(s_no,effort_date,Day,in_time,out_time,attendence,hours,project_id)
    #print m_file
    fields = []
    i = c2
    for k in range(0, 9,1):
        
        if k==  6:
           i = i + 1
        fields.append(str(worksheet.cell(r2,i).value))  
        i = i + 1
    #print fields
    #print str(worksheet.cell(22,2).value)
    m_list1 = []
    f = 0
    #m_list1 is the list of the dictionary in whch the parsed table is stored
    for m in m_file:
        
        dict1 = {fields[0]:m_file[f][0],
                 fields[1]:m_file[f][1],
                 fields[2]:m_file[f][2],
                 fields[3]:m_file[f][3],
                 fields[4]:m_file[f][4],
                 fields[5]:m_file[f][5],
                 fields[6]:m_file[f][6],
                 fields[7]:m_file[f][7]
                 
            }
        
        f = f + 1
        m_list1.append(dict1)
    #print m_list1
    dict3 = {}
    for i in range(r3,r4 + 1):
        c4 = get_value_column(worksheet,i,c3)
        dict3.update({
            str((worksheet.cell(i,c3).value)):str((worksheet.cell(i,c4)).value)
            })
    #print dict3
    file1 = open("sample_2.txt",'w')
    for dictionary in m_list1:
        for key,value in dictionary.iteritems():
            print key,value
            file1.write(key + ":"  +  value)
            file1.write("\t")
        file1.write("\n")
    for key,value in dict2.iteritems():
        print key,value
        file1.write(key + ":"  +  value)
        file1.write("\t")
    file1.write("\n")
    for key,value in dict3.iteritems():
        print key,value
        file1.write(key + ":"  +  value)
        file1.write("\t")
    file1.write("\n")
    file1.close()
    
    return m_list1,dict2,dict3
a,b,c = excel_extract("sample_2.xlsx")
print a,b,c
