#program for Worksheet in TIMESHEET PROCESS -NON IBM3.xlsx
#both the modules openpyxl and xlrd have some roles in the following code
import openpyxl
import xlrd
import re
#------------------------------------------------------------------------------#

#------------------------------------------------------------------------------#        
def get_value_column(worksheet,row_num ,column_num):
    for j in range(int(column_num) + 1,int(column_num) + 20,1):
        
        if str((worksheet.cell(int(row_num),j).value)) != '':
            return j
#------------------------------------------------------------------------------#

#------------------------------------------------------------------------------#                
#the function get_row_column returns the row and column number of a given value
def get_row_column(value,worksheet):
    for row in range(worksheet.nrows):
        for column in range(worksheet.ncols):
            if worksheet.cell(row,column).value == value:
                return row,column
#the function that extracts values from the xlsx
#------------------------------------------------------------------------------#

#------------------------------------------------------------------------------#        
def excel_extract(path):
    #xlrd.open_workbook(path,'cp1252') takes care of the unicodes
    workbook = xlrd.open_workbook(path,
                                  'cp1252')
    worksheet = workbook.sheet_by_index(0)
    r,c =  get_row_column("TIMESHEET",worksheet)
    r3,c3 = get_row_column("Name of the Resource *",worksheet)
    r4,c4 = get_row_column("Total Days in Month",worksheet)
    r1,c1 = get_row_column("Year",worksheet)
    dr,dc = get_row_column("DOJ",worksheet)
    days = ['Sat','Sun','Mon','Tue','Wed','Thu','Fri']
    for i in range(r,r4,1):
        if str((worksheet.cell(i,c).value)) in days:
            days_r = i
    #print days_r
    days_end_c = c
    i = c
    while str((worksheet.cell(days_r,i).value))!= '':
        i = i + 1
        days_end_c = i
    #print days_end_c

    #print dr,dc
    project_info = []

    project_info_values = []
    for i in range(r3,r4-1,1):
        project_info.append(str((worksheet.cell(i,c3).value)))
    days_array = []
    for i in range(c,days_end_c):
        days_array.append(str((worksheet.cell(days_r,i).value)))
    #print days_array
    dict1 = {}
    dict1_list = []
    for i in range(c,days_end_c,1):
        dict1 = {
            'days':str((worksheet.cell(days_r,i).value)),
            'dates':str((worksheet.cell(days_r + 1,i).value)),
            'present or absent or holiday':str((worksheet.cell(days_r + 2,i).value))
            }
            
        dict1_list.append(dict1)
    #print dict1_list
    #print project_info

    for i in range(r3,r4-1,1):
        project_info_values.append(str((worksheet.cell(i,get_value_column(worksheet,r3,c3)).value)))
        
    #print project_info_values
    time_data_type = []
    for i in range(r1,r3-5,1):
        time_data_type.append(str((worksheet.cell(i,c1).value)))
    #print time_data_type
    time_data_value = []
    for i in range(r1,r3-5,1):
        time_data_value.append(str((worksheet.cell(i,get_value_column(worksheet,r1,c1)).value)))
    #print time_data_value
    text_strings = []
    for i in range(r,r1-1,1):
        text_strings.append(str((worksheet.cell(i,c)).value))
    #print text_strings
    dict2 = {}
    i = 0
    for info in project_info:    
        dict2.update({project_info[i]:project_info_values[i]})
        i = i + 1
    #print dict2
    dict3 = {}
    for i in range(r4,r4+5,1):
        dict3.update({str((worksheet.cell(i,c4)).value):str((worksheet.cell(i,get_value_column(worksheet,r4,c4)).value))})
    #print dict3

    dc1 = get_value_column(worksheet,dr,dc)
    dc2 = get_value_column(worksheet,dr,dc1)
    dc3 = dc1 - dc     
    dict_joining_leave = {
                          str((worksheet.cell(dr,dc)).value):str((worksheet.cell(dr,dc1)).value),
                          str((worksheet.cell(dr,dc2)).value):
                          str((worksheet.cell(dr,dc3)).value)
                            }

    #print dict_joining_leave
    file1 = open("sample_5.txt",'w')
    for dictionary in dict1_list:
        for key,value in dictionary.iteritems():
            print key,value
            file1.write(key + ":"  +  value)
            file1.write("\t")
        file1.write("\n")
        file1.write("\n")
    for key,value in dict2.iteritems():
        print key,value
        file1.write(key + ":"  +  value)
        file1.write("\t")
    file1.write("\n")
    for key,value in dict3.iteritems():
        print key,value
        file1.write(key + ":"  +  value)
        file1.write("\t")
    file1.write("\n")
    for key,value in dict_joining_leave.iteritems():
        print key,value
        file1.write(key + ":"  +  value)
        file1.write("\t")
    file1.write("\n")
    file1.close()
    return dict1_list,dict2,dict3,dict_joining_leave
a,b,c,d = excel_extract("sample_5.xlsx")
print a,b,c,d    
