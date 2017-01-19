"Worksheet in TIMESHEET PROCESS -NON IBM2.xlsx"
#program for Worksheet in TIMESHEET PROCESS -NON IBM2.xlsx
#both the modules openpyxl and xlrd have some roles in the following code
import openpyxl
import xlrd
import re
#------------------------------------------------------------------------------#

#------------------------------------------------------------------------------
def get_value_column(worksheet,row_num ,column_num):
    for j in range(column_num + 1,column_num + 20,1):
        
        if str((worksheet.cell(row_num,j).value)) != '':
            return j
#------------------------------------------------------------------------------#

#------------------------------------------------------------------------------
        
#the function get_row_column returns the row and column number of a given value
def get_row_column(value,worksheet):
    for row in range(worksheet.nrows):
        for column in range(worksheet.ncols):
            if worksheet.cell(row,column).value == value:
                return row,column
#------------------------------------------------------------------------------#

#------------------------------------------------------------------------------
def excel_extract(path):
    #xlrd.open_workbook(path,'cp1252') takes care of the unicodes
    workbook = xlrd.open_workbook(path,
                                  'cp1252')
    worksheet = workbook.sheet_by_index(0)
    r1,c1 = get_row_column("Weekday",worksheet)
    #print r1,c1
    r2,c2 = get_row_column("Week 1",worksheet)
    r3,c3 = get_row_column("Week 2",worksheet)
    r4,c4 = get_row_column("Name of Project-in-charge",worksheet)
    diffrnc = (int(r3) - int(r2))
    r,c = get_row_column("Employee Name",worksheet)
    #print r4,c4

    table_dict = {}
    for i in range(r,r1 - 1):
        table_dict.update({
            str((worksheet.cell(i,c).value)): str((worksheet.cell(i,get_value_column(worksheet,r,c)).value))
            
            })
    #print table_dict
    c24 = get_value_column(worksheet,r4,c4)
    c24 = get_value_column(worksheet,r4,c24)
    footer_table_dict = {}
    for i in range(r4 - 1,r4 + 2):
        footer_table_dict.update({
            str((worksheet.cell(i,c4).value)): str((worksheet.cell(i,get_value_column(worksheet,r4,c4)).value)),
            str((worksheet.cell(i,c24).value)): str((worksheet.cell(i,c24 + 2).value))
            
            })
    #print footer_table_dict
    days = []
    for j in range(c1 + 1,c1 + 8,1):
        days.append(str((worksheet.cell(r1,j).value)))
    #print days
    for i in range(r2,r2+10,1):
         d = re.findall("Week",str((worksheet.cell(i,c2).value)))
    #print d
    i = r1
    dict1 = {}
    for i in range(r1 + 1,30,diffrnc):
        dict2 = {}
        for j in range(1,diffrnc,1):
            dict3 = {}
            for k in range(1,len(days),1):
                
                dict3.update({
                    days[k]: str((worksheet.cell(i + j ,k).value))
                })
                #print dict3
            dict2.update({
                str((worksheet.cell(i + j,c1).value)):dict3
                })
        dict1.update({
            str((worksheet.cell(i,c1).value)): dict2
            })
        #i = i + diffrnc
    #print dict1
    file1 = open("sample_4.txt",'w')
    for key,value in dict1.iteritems():        
            
            file1.write(key + ":" + "{" )
            for key,value in value.iteritems():
                file1.write(key + ":" + "{" )
                for key,value in value.iteritems():
                    file1.write(key + ":" + value)
                    #print key,value
                    file1.write("\t")
                file1.write("}")
                file1.write("\n")
            file1.write("}")
            file1.write("\n")
    file1.write("\n")
    for key,value in table_dict.iteritems():
        #print key,value
        file1.write(key + ":" + value)
        #file1.write(value)
        file1.write("\t")
    file1.write("\n")
    for key,value in footer_table_dict.iteritems():
        #print key,value
        file1.write(key + ":" + value)
        #file1.write(value)
        file1.write("\t")
    file1.write("\n")
    file1.close()
    return dict1,table_dict,footer_table_dict
    file1 = open("sample_42.txt",'w')
    file1.write(str(dict1))
    file1.close()
a,b,c = excel_extract("sample_4.xlsx")
print a,b,c
